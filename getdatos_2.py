import openpyxl
import pandas as pd

from openpyxl import load_workbook

# def read_excel_2d():
    

#     # Ruta del archivo Excel
#     archivo_excel = 'leyenda_v2.xlsx'
    
#     # Nombre de la hoja que deseas leer
#     nombre_hoja = 'NORMALIZADO'
    
#     # Rango de celdas (origen y destino) que deseas leer
#     origen = 'N4'
#     destino = 'BG49'
    
#     # Cargar el archivo Excel y la hoja deseada
#     book = load_workbook(archivo_excel)
#     hoja = book[nombre_hoja]
    
#     # Leer el rango de celdas específico y almacenar los datos en un DataFrame
#     df = pd.DataFrame(hoja[origen:destino])
    
#     # Convertir los datos en un array bidimensional
#     datos = df.values
    
#     # Verificar celdas vacías y mostrar mensaje
#     for fila in datos:
#         for valor in fila:
#             if pd.isna(valor):
#                 print("Se encontró una celda vacía.")
    
#     print(datos)
#     return datos



def read_excel_data_2d(filename, start_cell, end_cell, sheet_name):
    # Cargar el archivo de Excel
    wb = openpyxl.load_workbook(filename)
    
    # Obtener la hoja especificada por su nombre
    sheet = wb[sheet_name]
    
    # Crear un array bidimensional para almacenar los valores de las celdas
    values_2d = []
    
    # Recorrer las filas dentro del rango especificado
    for row in sheet[start_cell:end_cell]:
        row_values = []
        # Recorrer las celdas dentro de cada fila
        for cell in row:
            if cell.value is None:
                # Si la celda está vacía, imprimir un mensaje y agregar un valor nulo al array bidimensional
                print("Se encontró una celda vacía.")
                row_values.append(None)
            else:
                # Agregar el valor de la celda al array bidimensional
                row_values.append(cell.value)
        # Agregar la fila al array bidimensional
        values_2d.append(row_values)
    
    # Retornar el array bidimensional
    return values_2d


def getMatrizDistancia_Completa():
    filename = "leyenda_v2.xlsx"
    start_cell = "N4"
    end_cell = "BG49"
    sheet_name = "NORMALIZADO"
    data_2d = read_excel_data_2d(filename, start_cell, end_cell, sheet_name)
    # print(data_2d)
    
    matriz_distancia = data_2d
    
    #Pasar a enteros
    # matriz_adyacencia_enteros = [[int(elemento) for elemento in fila] for fila in matriz_adyacencia]
    # matriz_adyacencia = matriz_adyacencia_enteros
    
    return matriz_distancia

def getMatrizTiempo_Completa():
    filename = "leyenda_v2.xlsx"
    start_cell = "N56"
    end_cell = "BG101"
    sheet_name = "NORMALIZADO"
    data_2d = read_excel_data_2d(filename, start_cell, end_cell, sheet_name)
    # print(data_2d)
    
    matriz_tiempo = data_2d
    
    #Pasar a enteros
    # matriz_adyacencia_enteros = [[int(elemento) for elemento in fila] for fila in matriz_adyacencia]
    # matriz_adyacencia = matriz_adyacencia_enteros
    
    return matriz_tiempo

def getMatrizCalificacion_Completa():
    filename = "leyenda_v2.xlsx"
    start_cell = "N108"
    end_cell = "BG153"
    sheet_name = "NORMALIZADO"
    data_2d = read_excel_data_2d(filename, start_cell, end_cell, sheet_name)
    # print(data_2d)
    
    matriz_calificacion = data_2d
    
    #Pasar a enteros
    # matriz_adyacencia_enteros = [[int(elemento) for elemento in fila] for fila in matriz_adyacencia]
    # matriz_adyacencia = matriz_adyacencia_enteros
    
    return matriz_calificacion

# m = getMatrizDistancia_Completa()
# print(m)
# print("")
# print("")
# n = getMatrizTiempo_Completa()
# print(n)
# print("")
# print("")
# p = getMatrizCalificacion_Completa()
# print(p)
# print("")
# print("")
# print("Longitud: ", len(m[0]))
# print("Longitud: ", len(n[0]))
# print("Longitud: ", len(p[0]))
#read_excel_2d()